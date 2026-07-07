import io
import datetime
import pandas as pd
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

class ReportExporter:
    @staticmethod
    def to_csv(df: pd.DataFrame) -> bytes:
        """Converts DataFrame to CSV bytes."""
        output = io.StringIO()
        df.to_csv(output, index=False)
        return output.getvalue().encode("utf-8")

    @staticmethod
    def to_excel(df: pd.DataFrame, sheet_name: str = "Sheet1") -> bytes:
        """Converts DataFrame to styled Excel bytes."""
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            
            # Excel formatting
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            # Styling headers
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            thin_border = Border(
                left=Side(style='thin', color='D3D3D3'),
                right=Side(style='thin', color='D3D3D3'),
                top=Side(style='thin', color='D3D3D3'),
                bottom=Side(style='thin', color='D3D3D3')
            )
            
            for col_idx in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Aligning and formatting body
            body_alignment = Alignment(horizontal="left", vertical="center")
            for row in range(2, worksheet.max_row + 1):
                for col in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.alignment = body_alignment
                    cell.border = thin_border
            
            # Adjust column widths automatically
            for col in worksheet.columns:
                max_len = 0
                for cell in col:
                    val = str(cell.value or "")
                    if len(val) > max_len:
                        max_len = len(val)
                col_letter = col[0].column_letter
                worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
        return output.getvalue()

    @staticmethod
    def to_pdf(title: str, subtitle: str, df: pd.DataFrame) -> bytes:
        """Converts data to a high-fidelity PDF report using ReportLab."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=36,
            leftMargin=36,
            topMargin=36,
            bottomMargin=36
        )
        
        styles = getSampleStyleSheet()
        
        # Define Custom Styles
        title_style = ParagraphStyle(
            name="ReportTitle",
            parent=styles["Title"],
            fontName="Helvetica-Bold",
            fontSize=24,
            textColor=colors.HexColor("#1F497D"),
            alignment=0, # Left aligned
            spaceAfter=10
        )
        
        subtitle_style = ParagraphStyle(
            name="ReportSubtitle",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=11,
            textColor=colors.HexColor("#555555"),
            spaceAfter=20
        )
        
        th_style = ParagraphStyle(
            name="TableHeader",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=9,
            textColor=colors.white,
            alignment=1 # Centered
        )
        
        tb_style = ParagraphStyle(
            name="TableBody",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8,
            textColor=colors.HexColor("#333333")
        )

        elements = []
        
        # Header Info
        elements.append(Paragraph(title, title_style))
        elements.append(Paragraph(subtitle, subtitle_style))
        elements.append(Paragraph(f"Generated on: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", subtitle_style))
        elements.append(Spacer(1, 0.2 * inch))
        
        # Setup Table Data
        # Ensure column length is not too large for letter page size
        cols = list(df.columns)
        if len(cols) > 8:
            # truncate columns to fit on letter page
            cols = cols[:8]
            df = df[cols]
            
        data = [[Paragraph(col, th_style) for col in cols]]
        for idx, row in df.iterrows():
            row_cells = []
            for col in cols:
                val = str(row[col] if row[col] is not None else "")
                row_cells.append(Paragraph(val, tb_style))
            data.append(row_cells)
            
        # Draw Table
        table = Table(data, hAlign='LEFT', colWidths=[1.0 * inch] * len(cols))
        
        # Set Table Style
        t_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1F497D")),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#D3D3D3")),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ])
        
        # Alternating row colors
        for i in range(1, len(data)):
            if i % 2 == 0:
                t_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor("#F2F5F8"))
                
        table.setStyle(t_style)
        elements.append(table)
        
        # Build Document
        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()
